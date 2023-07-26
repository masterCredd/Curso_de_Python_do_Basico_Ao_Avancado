from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER/'exemplo_excel'/'workbook.xlsx'

workbook = Workbook()

# Nome para a planilha
sheet_name = 'Minha planilha'
# Criando a planilha
workbook.create_sheet(sheet_name, 0)
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Notas')

students = [
    # nome      idade nota
    ['João', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15,  8.8],
    ['Alberto', 16, 10],
]


# for i, students_row in enumerate(students, start=2):
#     for j, student_column in enumerate(students_row, start=1):
#         worksheet.cell(i, j, student_column)


for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)